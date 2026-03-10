"""Exportación de datos a Excel y PDF"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import datetime
from app.core.database import get_db
from app.api.dependencies import get_current_user
from app.schemas.auth import TokenData

router = APIRouter(prefix="/api/export", tags=["export"])

# ─── helpers ────────────────────────────────────────────────────────────────

def _get_rows(type_: str, seremiId: str | None, date_from: str | None, date_to: str | None, db: Session):
    """Devuelve (headers, rows) para el tipo solicitado."""
    if type_ == "contrataciones":
        from app.models.contratacion import Contratacion
        q = db.query(Contratacion)
        if seremiId:
            q = q.filter(Contratacion.seremiId == seremiId)
        rows = q.all()
        headers = ["ID", "SEREMI", "Nombre", "RUT", "Cargo", "Grado", "Tipo",
                   "Plaza", "Inicio", "Término", "Monto ($)", "Financiamiento",
                   "Estado", "VB Quien", "VB Fecha", "Motivo Rechazo", "Creado Por", "Creado En"]
        data = [[
            r.id, r.seremiId, r.nombre, r.rut, r.cargo, r.grado, r.tipo,
            r.esNuevo, r.inicio, r.termino, r.monto, r.financ,
            r.estado, r.vbQuien or "", r.vbFecha or "", r.motivoRechazo or "",
            r.creadoPor or "", r.creadoEn or ""
        ] for r in rows]
        return headers, data

    if type_ == "visitas":
        from app.models.visita import Visita
        q = db.query(Visita)
        if seremiId:
            q = q.filter(Visita.seremiId == seremiId)
        rows = q.all()
        headers = ["ID", "SEREMI", "Nombre", "Cargo/Institución", "Tipo", "Fecha", "Motivo", "Estado"]
        data = [[r.id, r.seremiId, r.nombre, getattr(r, 'cargo', ''), getattr(r, 'tipo', ''),
                 r.fecha, r.motivo, r.estado] for r in rows]
        return headers, data

    if type_ == "contactos":
        from app.models.contacto import Contacto
        q = db.query(Contacto)
        if seremiId:
            q = q.filter(Contacto.seremiId == seremiId)
        rows = q.all()
        headers = ["ID", "SEREMI", "Nombre", "Cargo", "Institución", "Email", "Teléfono", "Especialidad"]
        data = [[r.id, r.seremiId, r.nombre, r.cargo, r.institucion,
                 r.email, r.telefono, getattr(r, 'especialidad', '')] for r in rows]
        return headers, data

    if type_ == "prensa":
        from app.models.prensa import Prensa
        q = db.query(Prensa)
        if seremiId:
            q = q.filter(Prensa.seremiId == seremiId)
        rows = q.all()
        headers = ["ID", "SEREMI", "Título", "Medio", "Tipo", "Fecha", "Sentiment", "URL"]
        data = [[r.id, r.seremiId, r.titulo, r.medio, r.tipo,
                 r.fecha, getattr(r, 'sentiment', ''), getattr(r, 'url', '')] for r in rows]
        return headers, data

    if type_ == "proyectos":
        from app.models.proyecto import Proyecto
        q = db.query(Proyecto)
        if seremiId:
            q = q.filter(Proyecto.seremiId == seremiId)
        rows = q.all()
        headers = ["ID", "SEREMI", "Nombre", "Tipo", "Estado", "Avance %", "Inicio", "Fin", "Monto"]
        data = [[r.id, r.seremiId, r.nombre, getattr(r, 'tipo', ''),
                 r.estado, getattr(r, 'avance', ''), r.inicio, r.fin, getattr(r, 'monto', '')] for r in rows]
        return headers, data

    if type_ == "indicadores":
        from app.models.kpi import KPI
        q = db.query(KPI)
        if seremiId:
            q = q.filter(KPI.seremiId == seremiId)
        rows = q.all()
        headers = ["ID", "SEREMI", "Nombre", "Real", "Meta", "Unidad", "Periodo", "Descripción"]
        data = [[r.id, r.seremiId, r.nombre,
                 r.real, r.meta, r.unidad or '',
                 r.periodo or '', r.descripcion or ''] for r in rows]
        return headers, data

    if type_ == "foro":
        from app.models.foro import ForoTema
        q = db.query(ForoTema)
        rows = q.all()
        headers = ["ID", "Título", "Categoría", "Autor", "Fecha", "Posts"]
        data = [[r.id, r.titulo, getattr(r, 'categoria', ''), r.autorNombre,
                 r.creadoEn, getattr(r, 'numPosts', 0)] for r in rows]
        return headers, data

    raise HTTPException(status_code=400, detail=f"Tipo no soportado: {type_}")


# ─── Excel ───────────────────────────────────────────────────────────────────

def _build_excel(headers: list, data: list, title: str) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Header row
    header_fill = PatternFill("solid", fgColor="1E2D45")
    header_font = Font(name="Calibri", bold=True, color="E8EDF5", size=10)
    thin = Side(style="thin", color="2C3E50")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 24

    # Data rows
    alt_fill = PatternFill("solid", fgColor="F2F6FA")
    data_font = Font(name="Calibri", size=10)

    for row_idx, row in enumerate(data, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(data) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    # Freeze top row
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── PDF ─────────────────────────────────────────────────────────────────────

def _build_pdf(headers: list, data: list, title: str) -> BytesIO:
    from fpdf import FPDF, XPos, YPos

    class PDF(FPDF):
        def header(self):
            self.set_fill_color(21, 30, 46)
            self.rect(0, 0, 210, 20, "F")
            self.set_font("Helvetica", "B", 13)
            self.set_text_color(232, 237, 245)
            self.cell(0, 20, title, align="C")
            self.ln(6)

        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "", 8)
            self.set_text_color(120, 130, 145)
            self.cell(0, 10, f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Página {self.page_no()}", align="C")

    pdf = PDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()
    pdf.set_margins(10, 24, 10)

    # Column widths (distribute available width)
    page_w = pdf.w - 20
    col_w = min(page_w / len(headers), 38) if headers else 20
    widths = [col_w] * len(headers)

    # Header row
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(30, 45, 69)
    pdf.set_text_color(232, 237, 245)
    pdf.set_draw_color(50, 70, 100)
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 8, str(h)[:18], border=1, fill=True, align="C")
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", 7.5)
    for row_num, row in enumerate(data):
        if row_num % 2 == 0:
            pdf.set_fill_color(242, 246, 250)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 45, 69)
        for i, val in enumerate(row):
            pdf.cell(widths[i], 7, str(val or "")[:20], border=1, fill=True)
        pdf.ln()

    buf = BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return buf


# ─── Route ───────────────────────────────────────────────────────────────────

@router.get("")
def export_data(
    format: str = Query(..., description="pdf o excel"),
    type: str = Query(..., description="contrataciones|visitas|contactos|prensa|proyectos|indicadores|foro"),
    seremiId: str = Query(None),
    dateFrom: str = Query(None),
    dateTo: str = Query(None),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """Exportar datos a Excel o PDF"""
    # Restringir SEREMI a sus propios datos
    effective_seremi = seremiId
    if current_user.rol == "seremi":
        effective_seremi = current_user.seremiId

    headers, data = _get_rows(type, effective_seremi, dateFrom, dateTo, db)

    title_map = {
        "contrataciones": "Contrataciones",
        "visitas": "Visitas",
        "contactos": "Contactos",
        "prensa": "Prensa",
        "proyectos": "Proyectos SEIA",
        "indicadores": "Indicadores KPI",
        "foro": "Foro",
    }
    title = f"{title_map.get(type, type)} — {datetime.now().strftime('%d/%m/%Y')}"

    date_str = datetime.now().strftime("%Y-%m-%d")

    if format == "excel":
        buf = _build_excel(headers, data, title_map.get(type, type))
        filename = f"{type}-{date_str}.xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return StreamingResponse(
            buf,
            media_type=mime,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    elif format == "pdf":
        buf = _build_pdf(headers, data, title)
        filename = f"{type}-{date_str}.pdf"
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    else:
        raise HTTPException(status_code=400, detail="Formato inválido. Usa 'pdf' o 'excel'")
